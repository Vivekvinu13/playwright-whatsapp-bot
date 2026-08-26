import json
import random
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright, TimeoutError


# --------------------------------------------------
# FILES
# --------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent

CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
SESSION_DIR = BASE_DIR / "whatsapp_session"
SCREENSHOT_DIR = BASE_DIR / "screenshots"

today = datetime.now().strftime("%Y-%m-%d")

JSON_FILE = BASE_DIR / f"whatsapp_report_{today}.json"
EXCEL_FILE = BASE_DIR / f"whatsapp_report_{today}.xlsx"


# --------------------------------------------------
# RANDOM HUMAN-LIKE DELAY
# --------------------------------------------------

def human_delay():
    delay = random.uniform(2, 5)
    print(f"Waiting {delay:.1f} seconds...")
    time.sleep(delay)


# --------------------------------------------------
# READ CONTACTS
# --------------------------------------------------

def read_contacts():

    workbook = load_workbook(
        CONTACTS_FILE,
        data_only=True
    )

    sheet = workbook.active
    contacts = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        name, phone, message = row[:3]

        if not name or not phone:
            continue

        contacts.append({
            "name": str(name).strip(),
            "phone": str(phone).strip(),
            "message": str(message or "").strip()
        })

    workbook.close()
    return contacts


# --------------------------------------------------
# WAIT FOR WHATSAPP
# --------------------------------------------------

def wait_for_whatsapp(page):

    print("\nWaiting for WhatsApp Web...")

    page.wait_for_timeout(10000)

    selectors = [
        'div[contenteditable="true"]',
        '[aria-label*="Search"]',
        '[title*="Search"]',
        'button[aria-label*="Search"]'
    ]

    for selector in selectors:

        try:

            page.locator(selector).first.wait_for(
                state="visible",
                timeout=10000
            )

            print(
                f"WhatsApp Web is ready."
            )

            return True

        except TimeoutError:
            pass

    return False


# --------------------------------------------------
# SEARCH BOX
# --------------------------------------------------

def get_search_box(page):

    selectors = [
        'div[contenteditable="true"][data-tab="3"]',
        'div[contenteditable="true"][role="textbox"]',
        '[aria-label*="Search"]',
        '[title*="Search"]'
    ]

    for selector in selectors:

        box = page.locator(selector).first

        try:

            box.wait_for(
                state="visible",
                timeout=5000
            )

            return box

        except TimeoutError:
            pass

    raise TimeoutError(
        "WhatsApp search box not found."
    )


# --------------------------------------------------
# MESSAGE BOX
# --------------------------------------------------

def get_message_box(page):

    selectors = [
        'footer div[contenteditable="true"]',
        'div[contenteditable="true"][role="textbox"]'
    ]

    for selector in selectors:

        box = page.locator(selector).last

        try:

            box.wait_for(
                state="visible",
                timeout=5000
            )

            return box

        except TimeoutError:
            pass

    raise TimeoutError(
        "WhatsApp message box not found."
    )


# --------------------------------------------------
# SEARCH CONTACT
# --------------------------------------------------

def search_contact(page, name, phone):

    print(f"Searching for {name}...")

    search_box = get_search_box(page)

    # Search by name
    search_box.click()
    search_box.fill(name)

    page.wait_for_timeout(1500)

    try:

        result = page.get_by_text(
            name,
            exact=True
        ).first

        result.wait_for(
            state="visible",
            timeout=3000
        )

        result.click()

        human_delay()

        print(f"Opened chat for {name}.")
        return True

    except TimeoutError:
        pass

    # Search by phone
    print(f"Name not found. Searching {phone}...")

    search_box = get_search_box(page)

    search_box.click()
    search_box.fill("")
    search_box.fill(phone)

    page.wait_for_timeout(1500)

    for selector in [
        '[role="listitem"]',
        '[role="gridcell"]',
        'div[data-testid="cell-frame-container"]'
    ]:

        results = page.locator(selector)

        if results.count() > 0:

            try:

                results.first.click()

                human_delay()

                print(f"Opened chat for {phone}.")
                return True

            except Exception:
                pass

    # Direct chat fallback
    print("Trying direct chat...")

    clean_phone = re.sub(
        r"\D",
        "",
        phone
    )

    page.goto(
        f"https://web.whatsapp.com/send?phone={clean_phone}",
        wait_until="domcontentloaded"
    )

    page.wait_for_timeout(2000)

    try:

        get_message_box(page)

        human_delay()

        print(
            f"Opened direct chat for {phone}."
        )

        return True

    except TimeoutError:

        return False


# --------------------------------------------------
# LAST 3 INCOMING MESSAGES
# --------------------------------------------------

def get_last_three_messages(page):

    selectors = [
        "div.message-in span.selectable-text",
        "div.message-in div.copyable-text"
    ]

    messages = []

    for selector in selectors:

        locator = page.locator(selector)
        count = locator.count()

        for i in range(count):

            try:

                text = locator.nth(i).inner_text().strip()

                if text:
                    messages.append(text)

            except Exception:
                pass

    return messages[-3:]


# --------------------------------------------------
# SAVE REPORTS
# --------------------------------------------------

def save_reports(results):

    # JSON
    with open(
        JSON_FILE,
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            results,
            file,
            indent=4,
            ensure_ascii=False
        )

    # Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"

    sheet.append([
        "Name",
        "Phone",
        "Message",
        "Status",
        "Last 3 Messages",
        "Screenshot",
        "Error",
        "Timestamp"
    ])

    for result in results:

        sheet.append([
            result["name"],
            result["phone"],
            result["message"],
            result["status"],
            "\n".join(
                result["last_3_messages"]
            ),
            result["screenshot"],
            result["error"],
            result["timestamp"]
        ])

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 50
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 50
    sheet.column_dimensions["F"].width = 50

    workbook.save(EXCEL_FILE)


# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main():

    print("=" * 60)
    print("     WHATSAPP PLAYWRIGHT AUTOMATION")
    print("=" * 60)

    if not CONTACTS_FILE.exists():

        print(
            f"ERROR: {CONTACTS_FILE} not found."
        )

        return

    contacts = read_contacts()

    print(
        f"Contacts loaded: {len(contacts)}"
    )

    SCREENSHOT_DIR.mkdir(
        exist_ok=True
    )

    results = []

    with sync_playwright() as p:

        print("\nOpening WhatsApp Web...")

        context = p.chromium.launch_persistent_context(
            user_data_dir=str(SESSION_DIR),
            headless=False
        )

        page = (
            context.pages[0]
            if context.pages
            else context.new_page()
        )

        page.goto(
            "https://web.whatsapp.com",
            wait_until="domcontentloaded"
        )

        # --------------------------------------------------
        # LOGIN / READY CHECK
        # --------------------------------------------------

        ready = wait_for_whatsapp(page)

        if not ready:

            print(
                "\nWhatsApp interface was not detected."
            )

            print(
                "If the QR code is visible, scan it manually."
            )

            print(
                "Waiting another 30 seconds..."
            )

            page.wait_for_timeout(30000)

            ready = wait_for_whatsapp(page)

        if not ready:

            print(
                "WhatsApp still isn't ready."
            )

            print(
                "Close the browser and try again."
            )

            context.close()
            return

        # --------------------------------------------------
        # PROCESS CONTACTS
        # --------------------------------------------------

        for contact in contacts:

            name = contact["name"]
            phone = contact["phone"]
            template = contact["message"]

            message = (
                template.replace(
                    "{name}",
                    name
                )
                if template
                else
                f"Hello {name}, hope you're doing well!"
            )

            result = {
                "name": name,
                "phone": phone,
                "message": message,
                "status": "Failed",
                "last_3_messages": [],
                "screenshot": "",
                "error": "",
                "timestamp": datetime.now().isoformat()
            }

            print("\n" + "-" * 50)
            print(
                f"Processing: {name} - {phone}"
            )

            try:

                # Search and open contact
                if not search_contact(
                    page,
                    name,
                    phone
                ):

                    raise Exception(
                        f"Could not open contact: "
                        f"{name} / {phone}"
                    )

                # Find message box
                print(
                    "Finding message box..."
                )

                message_box = get_message_box(
                    page
                )

                message_box.wait_for(
                    state="visible",
                    timeout=15000
                )

                # Type message
                print(
                    "Typing personalized message..."
                )

                message_box.fill(
                    message
                )

                human_delay()

                # Send message
                print(
                    "Sending message..."
                )

                page.keyboard.press(
                    "Enter"
                )

                human_delay()

                # Screenshot
                print(
                    "Taking screenshot..."
                )

                sent_messages = page.locator(
                    "div.message-out"
                )

                if sent_messages.count() > 0:

                    screenshot_name = (
                        re.sub(
                            r"[^a-zA-Z0-9_-]",
                            "_",
                            name
                        )
                        + "_"
                        + today
                        + ".png"
                    )

                    screenshot_path = (
                        SCREENSHOT_DIR
                        / screenshot_name
                    )

                    sent_messages.last.screenshot(
                        path=str(
                            screenshot_path
                        )
                    )

                    result["screenshot"] = str(
                        screenshot_path
                    )

                # Extract last 3 messages
                print(
                    "Extracting last 3 messages..."
                )

                result["last_3_messages"] = (
                    get_last_three_messages(
                        page
                    )
                )

                result["status"] = "Sent"

                print(
                    f"✓ Message sent to {name}"
                )

            except Exception as error:

                result["error"] = str(
                    error
                )

                print(
                    f"✗ Failed for {name}: {error}"
                )

            results.append(
                result
            )

            # Required 2–5 second random delay
            human_delay()

        # --------------------------------------------------
        # SAVE REPORTS
        # --------------------------------------------------

        print(
            "\nSaving reports..."
        )

        save_reports(
            results
        )

        print(
            f"JSON report: {JSON_FILE}"
        )

        print(
            f"Excel report: {EXCEL_FILE}"
        )

        context.close()

    print(
        "\nAutomation completed."
    )


# --------------------------------------------------
# START
# --------------------------------------------------

if __name__ == "__main__":
    main()